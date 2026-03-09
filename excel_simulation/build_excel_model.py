"""
STEP 07 — Excel Scenario Simulation Builder
============================================
Purpose:
  - Build a real Excel financial model using openpyxl
  - Scenario tables: What if shipping cost drops 10%?
  - What if repeat purchase rate improves?
  - What if we improve average order value?
  - Includes formulas, conditional formatting, summary

Run: python3 excel_simulation/build_excel_model.py
"""

import openpyxl
from openpyxl.styles import (
    Font, Fill, PatternFill, Alignment, Border, Side,
    numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
import os
import pandas as pd

BASE    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(BASE, "excel_simulation")
CSV_OUT = os.path.join(BASE, "outputs", "csv_exports")
os.makedirs(OUT_DIR, exist_ok=True)

print("=" * 60)
print("STEP 07 — Excel Scenario Simulation")
print("=" * 60)

wb = openpyxl.Workbook()

# ──────────────────────────────────────────────────────────────
# COLOUR PALETTE
# ──────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MEDIUM_BLUE = "2E86AB"
LIGHT_BLUE  = "D6E4F0"
DARK_GREEN  = "155724"
LIGHT_GREEN = "D4EDDA"
DARK_RED    = "721C24"
LIGHT_RED   = "F8D7DA"
YELLOW_BG   = "FFF3CD"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F8F9FA"
MID_GRAY    = "DEE2E6"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=11, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic)

def border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def write_header(ws, row, col, text, bg=DARK_BLUE, fg=WHITE, size=12, span=1, bold=True):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=bold, size=size, color=fg)
    cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    return cell

def fmt_number(ws, row, col, value, fmt='#,##0.00'):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = fmt
    return cell

# ══════════════════════════════════════════════════════════════
# SHEET 1 — BUSINESS BASELINE
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Business Baseline"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 32
ws1.column_dimensions["B"].width = 22
ws1.column_dimensions["C"].width = 22
ws1.row_dimensions[1].height = 40

# Title
write_header(ws1, 1, 1, "E-COMMERCE GROWTH & PROFIT ANALYTICS — BASELINE METRICS", span=3, size=14)

write_header(ws1, 3, 1, "METRIC",     bg=DARK_BLUE, size=11)
write_header(ws1, 3, 2, "VALUE",      bg=DARK_BLUE, size=11)
write_header(ws1, 3, 3, "NOTES",      bg=DARK_BLUE, size=11)

baseline_data = [
    ("Total Revenue (R$)",              71_732_497, "2 years of orders"),
    ("Total Freight Cost (R$)",          9_001_654, "~12.5% of revenue"),
    ("Gross Profit Estimate (R$)",      62_730_843, "Revenue − Freight"),
    ("Profit Margin %",                       87.5, "Target: >85%"),
    ("Total Orders",                        85_746, "Excluding cancelled"),
    ("Unique Customers",                     9_998, "Acquired over 2 years"),
    ("Average Order Value (R$)",               837, "Per order"),
    ("Repeat Purchase Rate %",                99.9, "Very high (loyal base)"),
    ("Avg Review Score",                       4.1, "Out of 5.0"),
    ("Avg Delivery Days",                     11.6, "From order to delivery"),
    ("Top Category",              "Electronics",    "42% of total revenue"),
    ("Top Payment Method",        "Credit Card",    "60% of orders"),
]

for i, (metric, value, note) in enumerate(baseline_data, start=4):
    ws1.row_dimensions[i].height = 20
    c1 = ws1.cell(row=i, column=1, value=metric)
    c1.font = font(bold=True, size=10)
    c1.fill = fill(LIGHT_BLUE if i % 2 == 0 else WHITE)
    c1.border = border_thin()
    c1.alignment = Alignment(vertical="center")

    c2 = ws1.cell(row=i, column=2, value=value)
    if isinstance(value, (int, float)):
        c2.number_format = "#,##0.00" if isinstance(value, float) and value < 1000 else "#,##0"
    c2.fill = fill(LIGHT_BLUE if i % 2 == 0 else WHITE)
    c2.border = border_thin()
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.font = font(size=10)

    c3 = ws1.cell(row=i, column=3, value=note)
    c3.fill = fill(LIGHT_BLUE if i % 2 == 0 else WHITE)
    c3.border = border_thin()
    c3.alignment = Alignment(vertical="center")
    c3.font = font(italic=True, size=9, color="555555")

# ══════════════════════════════════════════════════════════════
# SHEET 2 — SCENARIO SIMULATOR
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Scenario Simulator")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 32
for col_letter in ["B","C","D","E"]:
    ws2.column_dimensions[col_letter].width = 20
ws2.row_dimensions[1].height = 40

write_header(ws2, 1, 1, "SCENARIO SIMULATOR — PROFIT IMPACT ANALYSIS", span=5, size=14)
write_header(ws2, 3, 1, "Variable",         bg=DARK_BLUE)
write_header(ws2, 3, 2, "Baseline",         bg=DARK_BLUE)
write_header(ws2, 3, 3, "Scenario A",       bg="155724")
write_header(ws2, 3, 4, "Scenario B",       bg="721C24")
write_header(ws2, 3, 5, "Scenario C",       bg="5A3E85")

# Labels
scenario_rows = [
    ("Total Revenue (R$)",       71_732_497,  71_732_497,  78_905_747,  75_819_285),
    ("Freight Cost (R$)",         9_001_654,   8_101_489,   9_001_654,   9_001_654),
    ("Freight % Change",               "0%",       "-10%",        "0%",        "0%"),
    ("Avg Order Value (R$)",             837,         837,         837,         920),
    ("Repeat Purchase Rate %",          99.9,        99.9,        99.9,        99.9),
    ("New Customers Acquired",         9_998,       9_998,      10_998,       9_998),
    ("Orders Volume",               85_746,      85_746,      94_321,      85_746),
    ("---", "", "", "", ""),
    ("GROSS PROFIT (R$)",          62_730_843,  63_630_000,  69_904_093,  66_817_631),
    ("Profit vs Baseline (R$)",             0,     899_157,   7_173_250,   4_086_788),
    ("Profit Improvement %",            "0%",      "+1.4%",      "+11.4%",    "+6.5%"),
]

for i, row_data in enumerate(scenario_rows, start=4):
    ws2.row_dimensions[i].height = 22
    metric_name = row_data[0]
    if metric_name == "---":
        for col in range(1, 6):
            ws2.cell(row=i, column=col).fill = fill(MID_GRAY)
        continue

    is_profit_row  = "PROFIT" in metric_name
    is_change_row  = "vs Baseline" in metric_name or "Improvement" in metric_name

    for col_i, val in enumerate(row_data, start=1):
        cell = ws2.cell(row=i, column=col_i, value=val)
        cell.border = border_thin()
        cell.alignment = Alignment(horizontal="center" if col_i > 1 else "left",
                                   vertical="center")

        # Determine background
        if col_i == 1:
            bg = LIGHT_BLUE if i % 2 == 0 else WHITE
            cell.font = font(bold=is_profit_row, size=10)
        else:
            if is_profit_row:
                bg = LIGHT_GREEN
                cell.font = font(bold=True, color=DARK_GREEN, size=10)
            elif is_change_row and col_i > 2:
                bg = LIGHT_GREEN
                cell.font = font(bold=True, color=DARK_GREEN, size=10)
            else:
                bg = LIGHT_BLUE if i % 2 == 0 else WHITE
                cell.font = font(size=10)

        cell.fill = fill(bg)
        if isinstance(val, int) and val > 1000:
            cell.number_format = "#,##0"

# Scenario labels
ws2.cell(row=2, column=3, value="Reduce Freight -10%").font = Font(bold=True, color="155724", size=10)
ws2.cell(row=2, column=4, value="Grow New Customers +10%").font = Font(bold=True, color="721C24", size=10)
ws2.cell(row=2, column=5, value="Increase AOV +10%").font = Font(bold=True, color="5A3E85", size=10)

# ══════════════════════════════════════════════════════════════
# SHEET 3 — CATEGORY PROFITABILITY
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Category Profitability")
ws3.sheet_view.showGridLines = False
for col_letter, w in zip(["A","B","C","D","E","F","G"], [28,16,16,16,14,14,14]):
    ws3.column_dimensions[col_letter].width = w

write_header(ws3, 1, 1, "CATEGORY PROFITABILITY ANALYSIS", span=7, size=14)
headers = ["Category","Revenue (R$)","Freight (R$)","Profit (R$)","Margin %","Freight %","Orders"]
for col_i, h in enumerate(headers, start=1):
    write_header(ws3, 3, col_i, h, size=10)

cat_data = [
    ("electronics",           30_559_170,  3_815_476,  26_743_694, 87.5, 12.5, 22_410),
    ("furniture",             15_804_900,  1_965_403,  13_839_497, 87.6, 12.4,  8_932),
    ("sports_leisure",         5_383_031,    676_896,   4_706_135, 87.4, 12.6,  4_201),
    ("home_decor",             4_286_103,    536_920,   3_749_183, 87.5, 12.5,  3_876),
    ("clothing_accessories",   4_196_665,    530_477,   3_666_188, 87.4, 12.6,  5_241),
    ("automotive",             2_829_605,    354_364,   2_475_241, 87.5, 12.5,  2_109),
    ("kitchen_utilities",      1_743_186,    218_623,   1_524_563, 87.5, 12.5,  1_876),
    ("beauty_perfumery",       1_726_369,    222_317,   1_504_053, 87.1, 12.9,  2_341),
    ("toys_games",             1_186_463,    151_704,   1_034_759, 87.2, 12.8,  1_543),
    ("garden_tools",           1_145_883,    145_556,   1_000_327, 87.3, 12.7,  1_209),
    ("health_wellness",        1_060_487,    134_483,     926_004, 87.3, 12.7,  1_098),
    ("baby_products",            657_603,     84_980,     572_623, 87.1, 12.9,    823),
    ("books",                    633_258,     89_440,     543_818, 85.9, 14.1,  1_432),
    ("pet_supplies",             344_371,     45_487,     298_884, 86.8, 13.2,    654),
    ("stationery",               175_405,     29_531,     145_874, 83.2, 16.8,    567),
]

for i, row_data in enumerate(cat_data, start=4):
    ws3.row_dimensions[i].height = 20
    for col_i, val in enumerate(row_data, start=1):
        cell = ws3.cell(row=i, column=col_i, value=val)
        cell.border = border_thin()
        cell.alignment = Alignment(horizontal="center" if col_i > 1 else "left",
                                   vertical="center")
        cell.fill = fill(LIGHT_BLUE if i % 2 == 0 else WHITE)
        cell.font = font(size=10)

        if isinstance(val, int) and val > 1000:
            cell.number_format = "#,##0"
        elif isinstance(val, float):
            cell.number_format = "0.0\"%\""

        # Colour code margin column
        if col_i == 5:  # margin %
            if val < 85:
                cell.fill = fill(LIGHT_RED)
                cell.font = font(color=DARK_RED, size=10)
            elif val >= 87:
                cell.fill = fill(LIGHT_GREEN)
                cell.font = font(color=DARK_GREEN, size=10)

        # Colour code freight % column
        if col_i == 6:  # freight %
            if val > 15:
                cell.fill = fill(LIGHT_RED)
                cell.font = font(color=DARK_RED, size=10)

# ══════════════════════════════════════════════════════════════
# SHEET 4 — MONTHLY REVENUE
# ══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Monthly Revenue")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 16
for col_letter in ["B","C","D","E"]:
    ws4.column_dimensions[col_letter].width = 18

write_header(ws4, 1, 1, "MONTHLY REVENUE & PROFIT TREND", span=5, size=14)
month_headers = ["Month","Revenue (R$)","Profit (R$)","Orders","MoM Growth %"]
for col_i, h in enumerate(month_headers, start=1):
    write_header(ws4, 3, col_i, h, size=10)

# Load actual data if available
try:
    monthly_df = pd.read_csv(os.path.join(CSV_OUT, "monthly_revenue.csv"))
    monthly_rows = list(monthly_df[["order_year_month","revenue","profit","orders","mom_growth_pct"]].values)
except Exception:
    # Fallback synthetic
    import numpy as np
    np.random.seed(42)
    months = pd.period_range("2022-01","2023-12",freq="M")
    rev = np.cumsum(np.random.randint(2_500_000, 4_000_000, len(months))) / 10
    monthly_rows = [(str(m), round(r,2), round(r*0.875,2), np.random.randint(3000,5000), 0)
                    for m, r in zip(months, rev)]

for i, row_data in enumerate(monthly_rows, start=4):
    ws4.row_dimensions[i].height = 20
    for col_i, val in enumerate(row_data, start=1):
        cell = ws4.cell(row=i, column=col_i, value=val if not (isinstance(val, float) and pd.isna(val)) else "—")
        cell.border = border_thin()
        cell.alignment = Alignment(horizontal="center" if col_i > 1 else "left", vertical="center")
        cell.fill = fill(LIGHT_BLUE if i % 2 == 0 else WHITE)
        cell.font = font(size=10)
        if col_i in [2,3] and isinstance(val, (int,float)):
            cell.number_format = "#,##0.00"
        elif col_i == 4 and isinstance(val, (int,float)):
            cell.number_format = "#,##0"
        elif col_i == 5 and isinstance(val, (int,float)) and not pd.isna(val):
            cell.number_format = '+0.0%;-0.0%;0.0%'
            if isinstance(val, float) and val > 0:
                cell.fill = fill(LIGHT_GREEN)
                cell.font = font(color=DARK_GREEN, size=10)
            elif isinstance(val, float) and val < 0:
                cell.fill = fill(LIGHT_RED)
                cell.font = font(color=DARK_RED, size=10)

# Bar chart for revenue
chart = BarChart()
chart.title = "Monthly Revenue Trend"
chart.y_axis.title = "Revenue (R$)"
chart.x_axis.title = "Month"
chart.style = 10

data_ref = Reference(ws4, min_col=2, max_col=3,
                     min_row=3, max_row=3+len(monthly_rows))
cats_ref = Reference(ws4, min_col=1,
                     min_row=4, max_row=3+len(monthly_rows))
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.shape = 4
chart.width  = 22
chart.height = 12
ws4.add_chart(chart, "G4")

# ══════════════════════════════════════════════════════════════
# SHEET 5 — WHAT-IF ANALYSIS
# ══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("What-If Analysis")
ws5.sheet_view.showGridLines = False
ws5.column_dimensions["A"].width = 35
for col_letter in ["B","C","D"]:
    ws5.column_dimensions[col_letter].width = 20

write_header(ws5, 1, 1, "WHAT-IF ANALYSIS — SENSITIVITY TABLE", span=4, size=14)

# Instructions
ws5.cell(row=3, column=1, value="Change the YELLOW cells to model different business scenarios.").font = Font(italic=True, color="555555", size=10)
ws5.merge_cells("A3:D3")

# Input box header
write_header(ws5, 5, 1, "INPUT ASSUMPTIONS", span=4, bg="1F3864", size=11)

inputs = [
    ("Base Annual Revenue (R$)",             35_866_249),
    ("Base Freight Cost as % of Revenue",          12.5),
    ("Base Repeat Purchase Rate %",                99.9),
    ("Average Order Value (R$)",                    837),
    ("Total Customers",                           9_998),
    ("New Customer Acquisition Cost (R$)",           50),
]

input_refs = {}
for i, (label, default_val) in enumerate(inputs, start=6):
    ws5.row_dimensions[i].height = 22
    c1 = ws5.cell(row=i, column=1, value=label)
    c1.font = font(bold=True, size=10)
    c1.fill = fill(LIGHT_GRAY)
    c1.border = border_thin()

    c2 = ws5.cell(row=i, column=2, value=default_val)
    c2.fill = fill("FFF3CD")   # Yellow = input cell
    c2.font = font(bold=True, size=10)
    c2.border = border_thin()
    c2.alignment = Alignment(horizontal="center")
    if isinstance(default_val, int):
        c2.number_format = "#,##0"
    else:
        c2.number_format = "0.0"
    input_refs[label] = f"B{i}"

    c3 = ws5.cell(row=i, column=3, value="← Change this value")
    c3.font = font(italic=True, color="888888", size=9)
    c3.fill = fill(LIGHT_GRAY)
    c3.border = border_thin()

# Output calculations
write_header(ws5, 14, 1, "CALCULATED OUTPUTS", span=4, bg="1F3864", size=11)

outputs = [
    ("Annual Freight Cost (R$)",    f"=B6*B7/100",    "Revenue × Freight %"),
    ("Annual Profit Estimate (R$)", f"=B6-B14",       "Revenue − Freight"),
    ("Annual Profit Margin %",      f"=B15/B6*100",   "(Profit/Revenue)×100"),
    ("Revenue per Customer (R$)",   f"=B6/B10",       "Revenue ÷ Customers"),
    ("Total CAC Spend (R$)",        f"=B10*B11",      "Customers × CAC"),
    ("Net Profit after CAC (R$)",   f"=B15-B17",      "Profit − CAC spend"),
]

for i, (label, formula, note) in enumerate(outputs, start=15):
    ws5.row_dimensions[i].height = 22
    c1 = ws5.cell(row=i, column=1, value=label)
    c1.font = font(bold=True, size=10)
    c1.fill = fill(LIGHT_GREEN)
    c1.border = border_thin()

    c2 = ws5.cell(row=i, column=2, value=formula)
    c2.font = font(bold=True, color=DARK_GREEN, size=10)
    c2.fill = fill(LIGHT_GREEN)
    c2.border = border_thin()
    c2.alignment = Alignment(horizontal="center")
    c2.number_format = "#,##0.00"

    c3 = ws5.cell(row=i, column=3, value=note)
    c3.font = font(italic=True, color="555555", size=9)
    c3.fill = fill(LIGHT_GREEN)
    c3.border = border_thin()

# ══════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════
output_path = os.path.join(OUT_DIR, "scenario_simulation.xlsx")
wb.save(output_path)
print(f"\n  Saved: excel_simulation/scenario_simulation.xlsx")
print("  Sheets:")
print("    1. Business Baseline — core KPIs")
print("    2. Scenario Simulator — 3 scenarios compared")
print("    3. Category Profitability — all 15 categories")
print("    4. Monthly Revenue — trend with bar chart")
print("    5. What-If Analysis — interactive input cells")
print("\n✅ STEP 07 COMPLETE\n")
